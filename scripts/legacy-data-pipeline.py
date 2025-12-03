#!/usr/bin/env python3
"""
Club 19 Sales OS - Legacy Data Ingestion Pipeline
Processes Hope and MC spreadsheets into normalized Xata-ready JSON
"""

import json
import uuid
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Any, Optional
import sys

# Check for required library
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import load_workbook

# File paths
HOPE_FILE = "/Users/olivertimmerman/Downloads/All HOPE_to Oct 2025..xlsx"
MC_FILE = "/Users/olivertimmerman/Downloads/All MC_to Oct 2025.xlsx"
OUTPUT_DIR = "/Users/olivertimmerman/Documents/Converso/Club-19-Sales-OS/data/legacy-import"

# Supplier normalization rules
SUPPLIER_AUTO_MERGE = {
    "Galaxy": "Galaxy VIC",
    "Galaxy Vic": "Galaxy VIC",
    "Galaxy VIC": "Galaxy VIC",
    "STOCK": "Stock (Internal)",
    "Stock": "Stock (Internal)",
    "In Stock": "Stock (Internal)",
    "Bags by Appointment": "Bags by Appointment",
    "Bags By Appointment": "Bags by Appointment",
    "BagsbyAppointment": "Bags by Appointment",
}

SUPPLIER_REQUIRES_REVIEW = ["L19 STOCK", "LOCAL", "TSUM"]


def parse_excel_date(value: Any) -> Optional[str]:
    """Parse Excel date to ISO format"""
    if value is None or value == "":
        return None

    # Already a datetime
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    # Excel serial number
    if isinstance(value, (int, float)):
        try:
            # Excel epoch is 1899-12-30
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            date_obj = excel_epoch + timedelta(days=int(value))
            return date_obj.strftime("%Y-%m-%d")
        except:
            return None

    # String date
    if isinstance(value, str):
        # Try common formats
        for fmt in ["%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                date_obj = datetime.strptime(value.strip(), fmt)
                return date_obj.strftime("%Y-%m-%d")
            except:
                continue

    return None


def parse_price(value: Any) -> float:
    """Parse price to float"""
    if value is None or value == "":
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        # Remove currency symbols and whitespace
        clean = value.strip().replace("£", "").replace("$", "").replace(",", "")
        try:
            return float(clean)
        except:
            return 0.0

    return 0.0


def clean_text(value: Any) -> str:
    """Clean text value"""
    if value is None or value == "":
        return ""

    return str(value).strip()


def title_case(value: str) -> str:
    """Convert to title case"""
    if not value:
        return ""
    return " ".join(word.capitalize() for word in value.split())


def load_hope_sheet() -> List[Dict[str, Any]]:
    """Load Hope's spreadsheet"""
    print(f"📖 Loading Hope's spreadsheet: {HOPE_FILE}")
    wb = load_workbook(HOPE_FILE, data_only=True)
    ws = wb.active

    rows = []
    headers = None

    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if idx == 1:
            # First row is headers
            headers = [str(cell).strip() if cell else f"col_{i}" for i, cell in enumerate(row)]
            print(f"   Headers: {headers}")
            continue

        if not any(row):  # Skip empty rows
            continue

        row_dict = {
            "date": parse_excel_date(row[0]) if len(row) > 0 else None,
            "invoice_number": clean_text(row[1]) if len(row) > 1 else "",
            "client_raw": clean_text(row[2]) if len(row) > 2 else "",
            "client_status_raw": clean_text(row[3]) if len(row) > 3 else "",
            "supplier_raw": clean_text(row[4]) if len(row) > 4 else "",
            "item_raw": clean_text(row[5]) if len(row) > 5 else "",
            "brand_raw": clean_text(row[6]) if len(row) > 6 else "",
            "category_raw": clean_text(row[7]) if len(row) > 7 else "",
            "buy_price_raw": parse_price(row[8]) if len(row) > 8 else 0.0,
            "sell_price_raw": parse_price(row[9]) if len(row) > 9 else 0.0,
            "margin_raw": parse_price(row[10]) if len(row) > 10 else 0.0,
            "source": "Hope",
            "row_number": idx,
        }

        rows.append(row_dict)

    print(f"   ✓ Loaded {len(rows)} rows from Hope")
    return rows


def load_mc_sheet() -> List[Dict[str, Any]]:
    """Load MC's spreadsheet"""
    print(f"📖 Loading MC's spreadsheet: {MC_FILE}")
    wb = load_workbook(MC_FILE, data_only=True)
    ws = wb.active

    rows = []
    headers = None

    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if idx == 1:
            # First row is headers (may be Unnamed)
            headers = [str(cell).strip() if cell else f"col_{i}" for i, cell in enumerate(row)]
            print(f"   Headers: {headers}")
            continue

        if not any(row):  # Skip empty rows
            continue

        row_dict = {
            "date": parse_excel_date(row[0]) if len(row) > 0 else None,
            "invoice_number": clean_text(row[1]) if len(row) > 1 else "",
            "client_raw": clean_text(row[2]) if len(row) > 2 else "",
            "client_status_raw": clean_text(row[3]) if len(row) > 3 else "",
            "supplier_raw": clean_text(row[4]) if len(row) > 4 else "",
            "item_raw": clean_text(row[5]) if len(row) > 5 else "",
            "brand_raw": clean_text(row[6]) if len(row) > 6 else "",
            "category_raw": clean_text(row[7]) if len(row) > 7 else "",
            "buy_price_raw": parse_price(row[8]) if len(row) > 8 else 0.0,
            "sell_price_raw": parse_price(row[9]) if len(row) > 9 else 0.0,
            "margin_raw": parse_price(row[10]) if len(row) > 10 else 0.0,
            "source": "MC",
            "row_number": idx,
        }

        rows.append(row_dict)

    print(f"   ✓ Loaded {len(rows)} rows from MC")
    return rows


def normalize_supplier(raw_supplier: str) -> Dict[str, Any]:
    """Normalize supplier name"""
    if not raw_supplier:
        return {"clean": "Unknown", "requires_review": True, "reason": "Empty supplier"}

    # Check auto-merge rules
    if raw_supplier in SUPPLIER_AUTO_MERGE:
        return {"clean": SUPPLIER_AUTO_MERGE[raw_supplier], "requires_review": False}

    # Check requires review
    if raw_supplier in SUPPLIER_REQUIRES_REVIEW:
        return {"clean": raw_supplier, "requires_review": True, "reason": "Ambiguous supplier name"}

    # Keep as-is
    return {"clean": raw_supplier, "requires_review": False}


def build_supplier_data(all_rows: List[Dict[str, Any]]) -> tuple:
    """Build supplier normalization map and audit table"""
    print("\n🔧 Building supplier data...")

    supplier_audit = {}
    supplier_map = {}

    for row in all_rows:
        raw = row["supplier_raw"]
        if not raw:
            raw = "Unknown"

        norm = normalize_supplier(raw)
        clean = norm["clean"]

        # Add to map
        supplier_map[raw] = norm

        # Add to audit
        if clean not in supplier_audit:
            supplier_audit[clean] = {
                "clean_name": clean,
                "raw_variants": set(),
                "sources": set(),
                "rows": [],
                "requires_review": norm.get("requires_review", False),
                "reason": norm.get("reason", ""),
            }

        supplier_audit[clean]["raw_variants"].add(raw)
        supplier_audit[clean]["sources"].add(row["source"])
        supplier_audit[clean]["rows"].append(row["row_number"])

    # Convert sets to lists for JSON
    for supplier in supplier_audit.values():
        supplier["raw_variants"] = sorted(list(supplier["raw_variants"]))
        supplier["sources"] = sorted(list(supplier["sources"]))

    print(f"   ✓ Found {len(supplier_audit)} unique suppliers")
    print(f"   ✓ {sum(1 for s in supplier_audit.values() if s['requires_review'])} require review")

    return supplier_map, supplier_audit


def build_client_data(all_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Build client audit table"""
    print("\n👥 Building client data...")

    client_audit = {}

    for row in all_rows:
        raw = row["client_raw"]
        if not raw:
            raw = "Unknown"

        # Normalize: lowercase + trim
        clean = raw.lower().strip()
        status = row["client_status_raw"]

        if clean not in client_audit:
            client_audit[clean] = {
                "client_clean": clean,
                "raw_variants": set(),
                "client_statuses": set(),
                "sources": set(),
                "first_seen": None,
                "last_seen": None,
                "trade_count": 0,
                "requires_review": False,
            }

        client_audit[clean]["raw_variants"].add(raw)
        client_audit[clean]["client_statuses"].add(status)
        client_audit[clean]["sources"].add(row["source"])
        client_audit[clean]["trade_count"] += 1

        # Track dates
        if row["date"]:
            if not client_audit[clean]["first_seen"] or row["date"] < client_audit[clean]["first_seen"]:
                client_audit[clean]["first_seen"] = row["date"]
            if not client_audit[clean]["last_seen"] or row["date"] > client_audit[clean]["last_seen"]:
                client_audit[clean]["last_seen"] = row["date"]

    # Check for conflicts (multiple statuses)
    for client in client_audit.values():
        client["raw_variants"] = sorted(list(client["raw_variants"]))
        client["client_statuses"] = sorted(list(client["client_statuses"]))
        client["sources"] = sorted(list(client["sources"]))

        if len(client["client_statuses"]) > 1:
            client["requires_review"] = True
            client["reason"] = f"Multiple statuses: {', '.join(client['client_statuses'])}"

    print(f"   ✓ Found {len(client_audit)} unique clients")
    print(f"   ✓ {sum(1 for c in client_audit.values() if c['requires_review'])} require review")

    return client_audit


def build_legacy_tables(all_rows: List[Dict[str, Any]], supplier_map: Dict, supplier_audit: Dict, client_audit: Dict) -> tuple:
    """Build final legacy tables for Xata import"""
    print("\n🗄️  Building legacy Xata tables...")

    # Build ID maps
    supplier_id_map = {s["clean_name"]: str(uuid.uuid4()) for s in supplier_audit.values()}
    client_id_map = {c["client_clean"]: str(uuid.uuid4()) for c in client_audit.values()}

    # Legacy Suppliers
    legacy_suppliers = []
    for supplier in supplier_audit.values():
        clean = supplier["clean_name"]
        supplier_id = supplier_id_map[clean]

        # Calculate first/last seen
        supplier_rows = [r for r in all_rows if normalize_supplier(r["supplier_raw"])["clean"] == clean]
        dates = [r["date"] for r in supplier_rows if r["date"]]

        legacy_suppliers.append({
            "id": supplier_id,
            "supplier_clean": clean,
            "raw_variants": supplier["raw_variants"],
            "requires_review": supplier["requires_review"],
            "reason": supplier.get("reason", ""),
            "first_seen": min(dates) if dates else None,
            "last_seen": max(dates) if dates else None,
            "trade_count": len(supplier_rows),
        })

    # Legacy Clients
    legacy_clients = []
    for client in client_audit.values():
        clean = client["client_clean"]
        client_id = client_id_map[clean]

        legacy_clients.append({
            "id": client_id,
            "client_clean": clean,
            "raw_variants": client["raw_variants"],
            "client_status": ", ".join(client["client_statuses"]),
            "first_seen": client["first_seen"],
            "last_seen": client["last_seen"],
            "trade_count": client["trade_count"],
            "requires_review": client.get("requires_review", False),
        })

    # Legacy Trades
    legacy_trades = []
    for row in all_rows:
        supplier_clean = normalize_supplier(row["supplier_raw"])["clean"]
        client_clean = row["client_raw"].lower().strip() if row["client_raw"] else "unknown"

        supplier_id = supplier_id_map.get(supplier_clean, "")
        client_id = client_id_map.get(client_clean, "")

        # Calculate margin if missing
        margin = row["margin_raw"]
        if not margin:
            margin = row["sell_price_raw"] - row["buy_price_raw"]

        # Infer category if empty
        category = row["category_raw"]
        if not category:
            if row["brand_raw"]:
                category = title_case(row["brand_raw"])
            elif row["item_raw"]:
                category = title_case(row["item_raw"].split()[0] if row["item_raw"] else "")

        legacy_trades.append({
            "id": str(uuid.uuid4()),
            "invoice_number": row["invoice_number"],
            "trade_date": row["date"],
            "raw_client": row["client_raw"],
            "raw_supplier": row["supplier_raw"],
            "client_id": client_id,
            "supplier_id": supplier_id,
            "item": row["item_raw"],
            "brand": row["brand_raw"],
            "category": title_case(category),
            "buy_price": row["buy_price_raw"],
            "sell_price": row["sell_price_raw"],
            "margin": margin,
            "source": row["source"],
            "raw_row": json.dumps(row),
        })

    print(f"   ✓ Built {len(legacy_suppliers)} legacy_suppliers")
    print(f"   ✓ Built {len(legacy_clients)} legacy_clients")
    print(f"   ✓ Built {len(legacy_trades)} legacy_trades")

    return legacy_suppliers, legacy_clients, legacy_trades


def build_live_supplier_seed(legacy_suppliers: List[Dict]) -> List[Dict]:
    """Build live supplier table seed"""
    print("\n🌱 Building live supplier seed...")

    suppliers_live = []
    for supplier in legacy_suppliers:
        suppliers_live.append({
            "id": supplier["id"],
            "supplier_name": supplier["supplier_clean"],
            "raw_variants": supplier["raw_variants"],
            "requires_review": supplier["requires_review"],
            "is_legacy": True,
            "created_from": "legacy-import-v1",
            "default_currency": "GBP",
            "supplier_type": None,
            "notes": None,
            "first_seen": supplier["first_seen"],
            "last_seen": supplier["last_seen"],
            "trade_count": supplier["trade_count"],
        })

    print(f"   ✓ Generated {len(suppliers_live)} supplier seeds")
    return suppliers_live


def main():
    """Main pipeline execution"""
    print("=" * 80)
    print("CLUB 19 SALES OS - LEGACY DATA INGESTION PIPELINE")
    print("=" * 80)

    # Create output directory
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # STEP 1: Load data
    print("\n📦 STEP 1: Loading and standardizing data...")
    hope_rows = load_hope_sheet()
    mc_rows = load_mc_sheet()
    all_rows = hope_rows + mc_rows
    print(f"   ✓ Total rows: {len(all_rows)}")

    # STEP 2: Build supplier data
    print("\n🏢 STEP 2: Extracting and normalizing suppliers...")
    supplier_map, supplier_audit = build_supplier_data(all_rows)

    # STEP 3: Build client data
    client_audit = build_client_data(all_rows)

    # STEP 4: Build legacy tables
    legacy_suppliers, legacy_clients, legacy_trades = build_legacy_tables(
        all_rows, supplier_map, supplier_audit, client_audit
    )

    # STEP 5: Build live supplier seed
    suppliers_live = build_live_supplier_seed(legacy_suppliers)

    # STEP 6: Write outputs
    print("\n💾 Writing output files...")

    outputs = {
        "supplier_normalisation_map.json": supplier_map,
        "supplier_audit_table.json": list(supplier_audit.values()),
        "legacy_suppliers.json": legacy_suppliers,
        "legacy_clients.json": legacy_clients,
        "legacy_trades.json": legacy_trades,
        "suppliers_live_seed.json": suppliers_live,
    }

    for filename, data in outputs.items():
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, "w") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"   ✓ {filename}")

    # STEP 7: Generate summary
    print("\n" + "=" * 80)
    print("📊 SUMMARY REPORT")
    print("=" * 80)
    print(f"Total Trades:              {len(legacy_trades)}")
    print(f"Total Suppliers:           {len(legacy_suppliers)}")
    print(f"  - Requires Review:       {sum(1 for s in legacy_suppliers if s['requires_review'])}")
    print(f"Total Clients:             {len(legacy_clients)}")
    print(f"  - Requires Review:       {sum(1 for c in legacy_clients if c['requires_review'])}")
    print(f"Hope Trades:               {len(hope_rows)}")
    print(f"MC Trades:                 {len(mc_rows)}")
    print(f"Date Range:                {min(r['date'] for r in all_rows if r['date'])} to {max(r['date'] for r in all_rows if r['date'])}")
    print(f"\nOutput Directory:          {OUTPUT_DIR}")
    print("=" * 80)
    print("✅ PIPELINE COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    main()
